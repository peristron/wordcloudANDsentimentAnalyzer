#     app_wordcloud_custom_v23.py
#  OPTIMIZED FOR DEPLOYMENT
# directory setup:
#   cd c:\users\oakhtar\documents\pyprojs_local  (replace name/path if needed)
# RUN -
#
#     streamlit run app_wordcloud_custom_v23.py
#
#
# multi-file word cloud generator (streaming + memory‑safe) with sentiment analysis
# purpose: generate per-file and combined word clouds + top unigrams (optional bigrams) from large csvs/xlsx/vtt
# input:
#   - csv (single or multi-column): choose to treat each line as raw text, or select specific columns to join.
#   - excel (.xlsx/.xlsm): select a sheet and specific columns to join.
#   - vtt (.vtt): parse transcript text directly, ignoring timestamps and cues.
# defaults:
#   - csv 'raw lines' mode is fastest if each line is one record of text.
#   - for multi-column csv/excel, stream only the selected columns for minimal memory usage.
# key options (sidebar):
#   - random seed, image size, colormap, background color, fonts.
#   - cleaning: strip html, unescape entities, remove urls; keep hyphens/apostrophes.
#   - stopwords: single words and phrases; optional common prepositions; integers; minimum word length.
#   - sentiment analysis: enable/disable, set thresholds and colors for positive/neutral/negative terms.
#   - performance: csv reader mode (raw lines, streaming csv columns, or pandas), csv chunk size (pandas), encoding.
# outputs:
#   - per-file and combined word clouds (png), top unigrams (csv), optional bigrams (csv).
# performance updates (v21-v23):
#   - [v21] expanded artifact cleaning to support zoom and teams transcripts (VTT timestamps, bracketed annotations).
#   - [v22] added direct support for .vtt file uploads with a dedicated, streaming VTT parser.
#   - [v23] added prominent data privacy disclaimers for safe public deployment.
# dependencies: python 3.10+, streamlit, pandas, matplotlib, wordcloud, openpyxl, nltk
# run: streamlit run this_file.py

import io
import re
import html
import gc
import time
import csv
import string
from collections import Counter
from typing import Dict, List, Tuple, Iterable, Optional, Callable

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from io import BytesIO
from wordcloud import WordCloud, STOPWORDS
from matplotlib import font_manager
from itertools import pairwise

# optional excel engine
try:
    import openpyxl
except Exception:
    openpyxl = None

# sentiment analysis engine
try:
    import nltk
    from nltk.sentiment.vader import SentimentIntensityAnalyzer
except ImportError:
    st.error("NLTK not found. Please run: `pip install nltk`")
    nltk = None
    SentimentIntensityAnalyzer = None


# precompiled patterns
HTML_TAG_RE = re.compile(r"<[^>]+>")
# Expanded regex for Slack, Zoom, and Teams chat/transcript artifacts
CHAT_ARTIFACT_RE = re.compile(
    r":\w+:"  # Emojis like :smile:
    r"|\b(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday|today|yesterday) at \d{1,2}:\d{2}\b"  # Slack Timestamps
    r"|\b\d+\s+repl(?:y|ies)\b"  # Slack Reply counts
    r"|\d{2}:\d{2}:\d{2}\.\d{3}\s+-->\s+\d{2}:\d{2}:\d{2}\.\d{3}"  # Zoom/VTT Timestamps
    r"|\[[^\]]+\]",  # Bracketed content from Teams/Zoom like [Music], [Applause], [00:15]
    flags=re.IGNORECASE
)


# ---------------------------
# utilities & setup
# ---------------------------

@st.cache_resource(show_spinner="Initializing sentiment analyzer...")
def setup_sentiment_analyzer():
    """
    Initializes the SentimentIntensityAnalyzer and downloads VADER lexicon if needed.
    Using @st.cache_resource ensures this runs only once.
    """
    if nltk is None:
        return None
    try:
        # check if already downloaded
        nltk.data.find('sentiment/vader_lexicon.zip')
    except LookupError:
        # download if not found
        nltk.download('vader_lexicon')
    return SentimentIntensityAnalyzer()


def prefer_index(options: List[str], preferred: List[str]) -> int:
    for name in preferred:
        if name in options:
            return options.index(name)
    return 0 if options else -1


@st.cache_data(show_spinner=False)
def list_system_fonts() -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for fe in font_manager.fontManager.ttflist:
        if fe.name not in mapping:
            mapping[fe.name] = fe.fname
    return dict(sorted(mapping.items(), key=lambda x: x[0].lower()))


def build_punct_translation(keep_hyphens: bool, keep_apostrophes: bool) -> dict:
    punct = string.punctuation
    if keep_hyphens:
        punct = punct.replace("-", "")
    if keep_apostrophes:
        punct = punct.replace("'", "")
    return str.maketrans("", "", punct)


def parse_user_stopwords(raw: str) -> Tuple[List[str], List[str]]:
    phrases, singles = [], []
    for item in [x.strip() for x in raw.split(",") if x.strip()]:
        if " " in item:
            phrases.append(item.lower())
        else:
            singles.append(item.lower())
    return phrases, singles


def default_prepositions() -> set:
    return {
        'about', 'above', 'across', 'after', 'against', 'along', 'among', 'around', 'at',
        'before', 'behind', 'below', 'beneath', 'beside', 'between', 'beyond', 'but', 'by',
        'concerning', 'despite', 'down', 'during', 'except', 'for', 'from', 'in', 'inside',
        'into', 'like', 'near', 'of', 'off', 'on', 'onto', 'out', 'outside', 'over', 'past',
        'regarding', 'since', 'through', 'throughout', 'to', 'toward', 'under', 'underneath',
        'until', 'up', 'upon', 'with', 'within', 'without'
    }


def build_phrase_pattern(phrases: List[str]) -> Optional[re.Pattern]:
    if not phrases:
        return None
    escaped = [re.escape(p) for p in phrases if p]
    if not escaped:
        return None
    return re.compile(rf"\b(?:{'|'.join(escaped)})\b", flags=re.IGNORECASE)


def estimate_row_count_from_bytes(file_bytes: bytes) -> int:
    """
    Approximate row count for progress (CSV/raw-lines).
    For CSV with quoted multi-line cells, this is an estimate, not exact.
    """
    if not file_bytes:
        return 0
    n = file_bytes.count(b"\n")
    if not file_bytes.endswith(b"\n"):
        n += 1
    return n


def format_duration(seconds: float) -> str:
    seconds = int(seconds)
    h, r = divmod(seconds, 3600)
    m, s = divmod(r, 60)
    if h > 0:
        return f"{h:d}:{m:02d}:{s:02d}"
    return f"{m:d}:{s:02d}"


def make_unique_header(raw_names: List[Optional[str]]) -> List[str]:
    """
    Ensure header names are strings and unique; fill empty with col_{i};
    deduplicate with suffix __k.
    """
    seen: Dict[str, int] = {}
    result: List[str] = []
    for i, nm in enumerate(raw_names):
        name = (str(nm).strip() if nm is not None else "")
        if not name:
            name = f"col_{i}"
        if name in seen:
            seen[name] += 1
            unique = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
            unique = name
        result.append(unique)
    return result


# ---------------------------
# row readers (raw lines / csv / excel / vtt)
# ---------------------------

def read_rows_raw_lines(file_bytes: bytes, encoding_choice: str = "auto") -> Iterable[str]:
    """
    Yields one line at a time with minimal memory usage.
    'auto' uses utf-8 (errors='replace'); 'latin-1' uses that explicitly.
    """
    def _iter_with_encoding(enc: str):
        bio = io.BytesIO(file_bytes)
        # errors='replace' avoids decode exceptions without large allocations
        with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline=None) as wrapper:
            for line in wrapper:
                yield line.rstrip("\r\n")

    if encoding_choice == "latin-1":
        yield from _iter_with_encoding("latin-1")
    else:
        yield from _iter_with_encoding("utf-8")


def read_rows_vtt(file_bytes: bytes, encoding_choice: str = "auto") -> Iterable[str]:
    """
    Parses a .vtt file, yielding only the spoken text lines.
    Skips header, timestamps, cue IDs, and blank lines.
    Attempts to strip speaker names like 'Name:' from the start of lines.
    """
    for line in read_rows_raw_lines(file_bytes, encoding_choice):
        line = line.strip()
        if not line:
            continue
        if line == "WEBVTT" or "-->" in line:
            continue
        if line.isdigit(): # Skip cue identifiers that are just numbers
            continue

        # Attempt to remove speaker name (e.g., "John Doe: Hello")
        if ":" in line:
            parts = line.split(":", 1)
            # Only strip if the part before the colon is relatively short and simple
            if len(parts) > 1 and len(parts[0]) < 30 and " " in parts[0]:
                yield parts[1].strip()
                continue

        yield line


def detect_csv_num_cols(
    file_bytes: bytes,
    encoding_choice: str = "auto",
    delimiter: str = ",",
) -> int:
    """
    Peek at first row to detect number of CSV columns.
    """
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        row = next(rdr, None)
        return len(row) if row is not None else 0


def get_csv_columns(
    file_bytes: bytes,
    encoding_choice: str = "auto",
    delimiter: str = ",",
    has_header: bool = True,
) -> List[str]:
    """
    Retrieve header names for CSV; if no header, synthesize as col_0..col_n-1 using first row length.
    """
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        first = next(rdr, None)
        if first is None:
            return []
        if has_header:
            return make_unique_header(first)
        else:
            return [f"col_{i}" for i in range(len(first))]


def iter_csv_selected_columns(
    file_bytes: bytes,
    encoding_choice: str,
    delimiter: str,
    has_header: bool,
    selected_columns: List[str],
    join_with: str = " ",
    drop_empty: bool = True,
) -> Iterable[str]:
    """
    Stream CSV rows, concatenate selected columns into a single text line.
    Handles quoting/newlines via Python csv module. Memory-safe.
    """
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        first = next(rdr, None)
        if first is None:
            return

        if has_header:
            header = make_unique_header(first)
            name_to_idx = {n: i for i, n in enumerate(header)}
            idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
        else:
            # synthesize names to align with UI selection
            header = [f"col_{i}" for i in range(len(first))]
            name_to_idx = {n: i for i, n in enumerate(header)}
            idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]

            # process the first row as data
            vals = [first[i] if i < len(first) else "" for i in idxs]
            if drop_empty:
                vals = [v for v in vals if v]
            yield join_with.join(str(v) for v in vals)

        # main loop
        for row in rdr:
            vals = [row[i] if i < len(row) else "" for i in idxs]
            if drop_empty:
                vals = [v for v in vals if v]
            yield join_with.join(str(v) for v in vals)


def get_excel_sheetnames(file_bytes: bytes) -> List[str]:
    if openpyxl is None:
        return []
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    return sheets


def get_excel_columns(
    file_bytes: bytes,
    sheet_name: str,
    has_header: bool = True,
) -> List[str]:
    """
    Read first row of the sheet to produce header names. If no header, synthesize names
    based on the first row's length (col_0..col_n-1).
    """
    if openpyxl is None:
        return []
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name]
    gen = ws.iter_rows(values_only=True, min_row=1, max_row=1)
    first = next(gen, None)
    wb.close()
    if first is None:
        return []
    if has_header:
        return make_unique_header(list(first))
    else:
        return [f"col_{i}" for i in range(len(first))]


def excel_estimate_rows(
    file_bytes: bytes,
    sheet_name: str,
    has_header: bool = True,
) -> int:
    if openpyxl is None:
        return 0
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name]
    total = ws.max_row or 0
    wb.close()
    if has_header and total > 0:
        total -= 1
    return max(total, 0)


def iter_excel_selected_columns(
    file_bytes: bytes,
    sheet_name: str,
    has_header: bool,
    selected_columns: List[str],
    join_with: str = " ",
    drop_empty: bool = True,
) -> Iterable[str]:
    """
    Stream Excel (.xlsx/.xlsm) rows via openpyxl in read_only mode, yielding a concatenation
    of selected columns per row. Memory-safe.
    """
    if openpyxl is None:
        return
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    first = next(rows_iter, None)
    if first is None:
        wb.close()
        return

    if has_header:
        header = make_unique_header(list(first))
        name_to_idx = {n: i for i, n in enumerate(header)}
        idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
    else:
        header = [f"col_{i}" for i in range(len(first))]
        name_to_idx = {n: i for i, n in enumerate(header)}
        idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]

        # process the first row as data
        vals = [first[i] if i < len(first) else "" for i in idxs]
        if drop_empty:
            vals = [v for v in vals if v]
        yield join_with.join("" if v is None else str(v) for v in vals)

    for row in rows_iter:
        vals = [row[i] if (row is not None and i < len(row)) else "" for i in idxs]
        if drop_empty:
            vals = [v for v in vals if v]
        yield join_with.join("" if v is None else str(v) for v in vals)

    wb.close()


# ---------------------------
# pandas fallback (single/multi-col csv with chunks)
# ---------------------------

def read_rows_csv_pandas(
    file_bytes: bytes,
    encoding_choice: str = "auto",
    chunksize: int = 10_000,
    header: bool = False,
    usecols: Optional[List[int]] = None,
    join_with: str = " ",
) -> Iterable[str]:
    """
    Pandas-based CSV chunk reader. Heavier than csv module; use when needed.
    If header=True, header row is treated by pandas; usecols can be names or indices.
    If header=False, usecols must be integer indices (0-based).
    """
    def _reader(enc: str):
        buf = io.BytesIO(file_bytes)
        pd_header = 0 if header else None
        return pd.read_csv(
            buf,
            header=pd_header,
            dtype=str,
            usecols=usecols,
            chunksize=chunksize,
            encoding=enc,
            engine="python",
            on_bad_lines="skip",
            memory_map=False,
        )

    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    try:
        reader = _reader(enc)
        for chunk in reader:
            if isinstance(chunk, pd.DataFrame):
                if chunk.empty:
                    continue
                # ensure strings, replace NaNs, then build joined rows
                chunk = chunk.fillna("")
                arr = chunk.to_numpy(dtype=str, copy=False)
                for row in arr:
                    yield join_with.join([v for v in row if v])
    except Exception:
        # fallback to latin-1 if utf-8 failed
        try:
            reader = _reader("latin-1")
            for chunk in reader:
                if isinstance(chunk, pd.DataFrame):
                    if chunk.empty:
                        continue
                    chunk = chunk.fillna("")
                    arr = chunk.to_numpy(dtype=str, copy=False)
                    for row in arr:
                        yield join_with.join([v for v in row if v])
        except Exception:
            # ultimate fallback: raw lines
            for line in read_rows_raw_lines(file_bytes, encoding_choice="latin-1"):
                yield line


# ---------------------------
# core processing
# ---------------------------

def is_url_token(tok: str) -> bool:
    t = tok.strip("()[]{}<>,.;:'\"!?").lower()
    if not t:
        return False
    return ("://" in t) or t.startswith("www.")


def process_rows_iter(
    rows_iter: Iterable[str],
    remove_chat_artifacts: bool,
    remove_html_tags: bool,
    unescape_entities: bool,
    remove_urls: bool,
    keep_hyphens: bool,
    keep_apostrophes: bool,
    user_phrase_stopwords: Tuple[str, ...],
    user_single_stopwords: Tuple[str, ...],
    add_preps: bool,
    drop_integers: bool,
    min_word_len: int,
    compute_bigrams: bool = False,
    progress_cb: Optional[Callable[[int], None]] = None,
    update_every: int = 2_000,
) -> Dict:
    start_time = time.perf_counter()
    stopwords = set(STOPWORDS)
    stopwords.update(user_single_stopwords)
    if add_preps:
        stopwords.update(default_prepositions())

    translate_map = build_punct_translation(keep_hyphens=keep_hyphens, keep_apostrophes=keep_apostrophes)
    phrase_pattern = build_phrase_pattern(list(user_phrase_stopwords))
    counts = Counter()
    bigram_counts = Counter() if compute_bigrams else None
    total_rows = 0

    # localize flags for speed
    _remove_chat = remove_chat_artifacts
    _remove_html = remove_html_tags
    _unescape = unescape_entities
    _remove_urls = remove_urls
    _min_len, _drop_int, _stopwords = min_word_len, drop_integers, stopwords
    _is_url, _trans, _ppat = is_url_token, translate_map, phrase_pattern

    for line in rows_iter:
        total_rows += 1
        text = line if isinstance(line, str) else ("" if line is None else str(line))

        if _remove_chat:
            text = CHAT_ARTIFACT_RE.sub(" ", text)

        if _remove_html:
            text = HTML_TAG_RE.sub(" ", text)

        if _unescape:
            try:
                text = html.unescape(text)
            except MemoryError:
                pass
        
        text = text.lower()
        if _ppat:
            text = _ppat.sub(" ", text)

        filtered_tokens_line: List[str] = []
        for t in text.split():
            if _remove_urls and _is_url(t): continue
            t = t.translate(_trans)
            if not t or len(t) < _min_len or (_drop_int and t.isdigit()) or t in _stopwords:
                continue
            filtered_tokens_line.append(t)

        if filtered_tokens_line:
            counts.update(filtered_tokens_line)
            if compute_bigrams and len(filtered_tokens_line) > 1:
                bigram_counts.update(tuple(bg) for bg in pairwise(filtered_tokens_line))

        if progress_cb and (total_rows % update_every == 0):
            progress_cb(total_rows)

    if progress_cb:
        progress_cb(total_rows)
    elapsed = time.perf_counter() - start_time
    return {"counts": counts, "bigrams": bigram_counts or Counter(), "rows": total_rows, "elapsed": elapsed}


# ---------------------------
# sentiment analysis logic
# ---------------------------

@st.cache_data(show_spinner="Analyzing term sentiment...")
def get_sentiments(_analyzer, terms: Tuple[str, ...]) -> Dict[str, float]:
    if not _analyzer or not terms: return {}
    return {term: _analyzer.polarity_scores(term)['compound'] for term in terms}

def create_sentiment_color_func(sentiments: Dict[str, float], pos_color: str, neg_color: str, neu_color: str, pos_threshold: float, neg_threshold: float):
    def color_func(word, font_size, position, orientation, random_state=None, **kwargs):
        score = sentiments.get(word, 0.0)
        if score >= pos_threshold: return pos_color
        elif score <= neg_threshold: return neg_color
        else: return neu_color
    return color_func

def get_sentiment_category(score: float, pos_threshold: float, neg_threshold: float) -> str:
    if score >= pos_threshold: return "Positive"
    if score <= neg_threshold: return "Negative"
    return "Neutral"

# ---------------------------
# visualization
# ---------------------------

def build_wordcloud_figure_from_counts(counts: Counter, max_words: int, width: int, height: int, bg_color: str, colormap: str, font_path: Optional[str], random_state: int, color_func: Optional[Callable] = None):
    limited = dict(counts.most_common(max_words))
    wc = WordCloud(width=width, height=height, background_color=bg_color, colormap=colormap, font_path=font_path, random_state=random_state, color_func=color_func, collocations=False, normalize_plurals=False).generate_from_frequencies(limited)
    fig_w, fig_h = max(6.0, width / 100.0), max(3.0, height / 100.0)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=100)
    ax.imshow(wc, interpolation="bilinear")
    ax.axis("off")
    plt.tight_layout()
    return fig, wc

def fig_to_png_bytes(fig: plt.Figure) -> BytesIO:
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.1)
    buf.seek(0)
    return buf

# ---------------------------
# streamlit app
# ---------------------------

st.set_page_config(page_title="multi-file word cloud generator", layout="wide")
st.title("🧠 multi-file word cloud generator (streaming + sentiment)[v23]")

# --- V23: ADDED PROMINENT DATA PRIVACY DISCLAIMER ---
st.warning("""
**⚠️ Data Privacy & Security Notice**

UPSHOT: When in doubt, pre-sanitize your data. It's easy to do, and there are several simple ways to do this. This is a public app running on Streamlit Community Cloud. **Do not upload files containing sensitive, private, or proprietary information.** All data you upload is processed on public servers. Please ensure your data is anonymized or masked *before* uploading.
""")

analyzer = setup_sentiment_analyzer()

st.sidebar.info("""
**Performance Tip:** This app streams data for memory safety. It performs well on files up to a few hundred MB. For very large files (>400MB), processing may be slow or fail due to server limits.
""")
uploaded_files = st.sidebar.file_uploader(
    "upload files (csv, xlsx/xlsm, vtt)",
    type=["csv", "xlsx", "xlsm", "vtt"],
    accept_multiple_files=True,
    help="Warning: Do not upload sensitive data. This app runs on a public server." # V23: Added help text for privacy
)

st.sidebar.markdown("### 🎨 appearance")
bg_color = st.sidebar.color_picker("background color", value="#ffffff")
colormap = st.sidebar.selectbox("colormap (for neutral words)", options=["viridis", "plasma", "inferno", "magma", "cividis", "tab10", "tab20", "Dark2", "Set3", "rainbow", "cubehelix", "prism", "Blues", "Greens", "Oranges", "Reds", "Purples", "Greys"], index=0)
width = st.sidebar.slider("image width (px)", 600, 2400, 1200, 100)
height = st.sidebar.slider("image height (px)", 300, 1400, 600, 50)
random_state = st.sidebar.number_input("random seed", 0, value=42, step=1, help="fixes the random layout.")

st.sidebar.markdown("### 🔬 sentiment analysis")
enable_sentiment = st.sidebar.checkbox("enable sentiment analysis", value=False, help="colours words by sentiment. requires `nltk` library")
if enable_sentiment and analyzer is None:
    st.sidebar.error("NLTK is not installed. Sentiment analysis is disabled.")
    enable_sentiment = False
pos_threshold, neg_threshold, pos_color, neu_color, neg_color = 0.05, -0.05, '#2ca02c', '#808080', '#d62728'
if enable_sentiment:
    c1, c2 = st.sidebar.columns(2)
    with c1: pos_threshold = st.slider("positive threshold", 0.0, 1.0, 0.05, 0.01)
    with c2: neg_threshold = st.slider("negative threshold", -1.0, 0.0, -0.05, 0.01)
    c1, c2, c3 = st.sidebar.columns(3)
    with c1: pos_color = st.color_picker("positive color", value=pos_color)
    with c2: neu_color = st.color_picker("neutral color", value=neu_color)
    with c3: neg_color = st.color_picker("negative color", value=neg_color)

# sidebar: cleaning options
st.sidebar.markdown("### 🧹 cleaning")
remove_chat_artifacts = st.sidebar.checkbox("remove chat/transcript artifacts", value=True, help="removes artifacts from slack, zoom, ms teams, etc. (emojis, timestamps, reply counts, [Music])")
remove_html_tags = st.sidebar.checkbox("strip html tags", value=True)
unescape_entities = st.sidebar.checkbox("unescape html entities (&amp; → &)", value=True)
remove_urls = st.sidebar.checkbox("remove urls", value=True)
keep_hyphens = st.sidebar.checkbox("keep hyphens (e.g., 'real-time')", value=False)
keep_apostrophes = st.sidebar.checkbox("keep apostrophes (e.g., don't)", value=False)

# sidebar: stopwords
st.sidebar.markdown("### 🛑 stopwords")
user_input = st.sidebar.text_area("words/phrases to ignore (comma-separated)", value="snhu, snhutest, snhudev, firstname.lastname, jane doe", help="phrases and usernames are supported (e.g., 'machine learning', 'firstname.lastname', 'jane doe').")
user_phrase_stopwords, user_single_stopwords = parse_user_stopwords(user_input)
add_preps = st.sidebar.checkbox("remove common prepositions", value=True)
drop_integers = st.sidebar.checkbox("remove integer-only tokens", value=True)
min_word_len = st.sidebar.slider("minimum word length", 1, 10, 2, help="drops tokens shorter than this length.")

# sidebar UI
st.sidebar.markdown("### 📊 tables")
top_n = st.sidebar.number_input("number of top terms to display", 5, 10000, 20)
st.sidebar.markdown("### 🔤 fonts")
font_map, font_names = list_system_fonts(), list(list_system_fonts().keys())
preferred_defaults = ["cmtt10", "cmr10", "Arial", "DejaVu Sans", "Helvetica", "Times New Roman", "Verdana"]
default_font_index = prefer_index(font_names, preferred_defaults)
combined_font_name = st.sidebar.selectbox("font for combined word cloud", font_names or ["(default)"], max(default_font_index, 0))
combined_font_path = font_map.get(combined_font_name) if font_names else None
with st.sidebar.expander("⚙️ performance options", expanded=False):
    encoding_choice = st.selectbox("file encoding (CSV)", ["auto (utf-8)", "latin-1"])
    st.caption("set CSV delimiter per file in its 'input options' expander.")
    chunksize = st.number_input("csv chunk size (rows) [pandas fallback]", 1_000, 100_000, 10_000, 1_000)
    compute_bigrams = st.checkbox("compute bigram table (more memory)", value=False)

# ---------------------------
# main processing loop
# ---------------------------
combined_counts, combined_bigrams, file_results = Counter(), Counter(), []
if uploaded_files:
    st.subheader("📄 per-file word clouds")
    overall_bar, overall_status = st.progress(0), st.empty()
    use_combined_option = "use combined font"
    total_files = len(uploaded_files)

    for idx, file in enumerate(uploaded_files):
        file_bytes, fname, lower = file.getvalue(), file.name, file.name.lower()
        is_csv = lower.endswith(".csv")
        is_xlsx = lower.endswith((".xlsx", ".xlsm"))
        is_vtt = lower.endswith(".vtt")

        if font_names:
            per_file_font_choice = st.sidebar.selectbox(f"font for {fname}", [use_combined_option] + font_names, 0, key=f"font_{idx}")
            per_file_font_path = combined_font_path if per_file_font_choice == use_combined_option else font_map.get(per_file_font_choice)
        else: per_file_font_choice, per_file_font_path = "(default)", None
        
        with st.expander(f"🧩 input options for: {fname}", expanded=False):
            if is_vtt:
                st.info("`.vtt` transcript file detected. Parsing speech content directly.")
            elif is_csv:
                try: inferred_cols = detect_csv_num_cols(file_bytes, encoding_choice, delimiter=",")
                except Exception: inferred_cols = 1
                default_mode = "csv columns (streaming, recommended)" if inferred_cols > 1 else "raw lines (treat each line as full text)"
                read_mode = st.radio("csv read mode", ["raw lines (treat each line as full text)", "csv columns (streaming, recommended)", "csv columns (pandas fallback)"], index=["raw lines (treat each line as full text)", "csv columns (streaming, recommended)", "csv columns (pandas fallback)"].index(default_mode), key=f"csv_mode_{idx}")
                delim_choice = st.selectbox("csv delimiter", [", (comma)", "tab (\\t)", "; (semicolon)", "| (pipe)"], 0, key=f"csv_delim_{idx}")
                delimiter = {", (comma)": ",", "tab (\\t)": "\t", "; (semicolon)": ";", "| (pipe)": "|"}[delim_choice]
                has_header = st.checkbox("first row is header", value=True if inferred_cols > 1 else False, key=f"csv_has_header_{idx}")
                selected_cols: List[str] = []
                join_with = " "
                if read_mode != "raw lines (treat each line as full text)":
                    try: col_names = get_csv_columns(file_bytes, encoding_choice, delimiter, has_header)
                    except Exception: col_names = []
                    default_cols = [col_names[0]] if col_names else []
                    selected_cols = st.multiselect("columns to use", col_names, default_cols, key=f"csv_cols_{idx}")
                    join_with = st.text_input("join selected columns with", " ", key=f"csv_join_{idx}", help=( "Separator inserted between values from the selected columns when combining them into one text row. " "Examples: space ' ' → 'Title text Comment text'; ' - ' → 'Title text - Comment text'; " "' | ' → 'Title text | Comment text'. Tip: Avoid empty '' to prevent words from merging. " "Note: The separator also influences bigrams at column boundaries."))
                    _ex_left, _ex_right = (selected_cols[0] if selected_cols else "Title"), (selected_cols[1] if len(selected_cols) > 1 else "Comment")
                    st.caption(f"Example with two columns: '{_ex_left} text{join_with}{_ex_right} text'")
            elif is_xlsx:
                if openpyxl is None: st.error("openpyxl not installed.")
                sheet_name, selected_cols, has_header, join_with = None, [], True, " "
                if openpyxl is not None:
                    try: sheets = get_excel_sheetnames(file_bytes)
                    except Exception as e: sheets, _ = [], st.error(f"failed to open workbook: {e}")
                    sheet_name = st.selectbox("sheet", sheets or ["(no sheets)"], 0, key=f"xlsx_sheet_{idx}")
                    has_header = st.checkbox("first row is header", True, key=f"xlsx_has_header_{idx}")
                    try: col_names = get_excel_columns(file_bytes, sheet_name, has_header) if sheets else []
                    except Exception: col_names = []
                    default_cols = [col_names[0]] if col_names else []
                    selected_cols = st.multiselect("columns to use", col_names, default_cols, key=f"xlsx_cols_{idx}")
                    join_with = st.text_input("join selected columns with", " ", key=f"xlsx_join_{idx}", help=( "Separator inserted between values from the selected columns when combining them into one text row. " "Examples: space ' ' → 'Title text Comment text'; ' - ' → 'Title text - Comment text'; " "' | ' → 'Title text | Comment text'. Tip: Avoid empty '' to prevent words from merging. " "Note: The separator also influences bigrams at column boundaries."))
                    _ex_left, _ex_right = (selected_cols[0] if selected_cols else "Title"), (selected_cols[1] if len(selected_cols) > 1 else "Comment")
                    st.caption(f"Example with two columns: '{_ex_left} text{join_with}{_ex_right} text'")
            else: st.info("unrecognized extension. defaulting to raw lines.")
        
        container = st.container()
        with container:
            st.markdown(f"#### {fname}")
            per_file_bar, per_file_status = st.progress(0), st.empty()
        
        rows_iter: Iterable[str] = iter([])
        approx_rows, update_every = 0, 2000
        
        if is_vtt:
            rows_iter = read_rows_vtt(file_bytes, encoding_choice=("latin-1" if encoding_choice == "latin-1" else "auto"))
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        elif is_csv:
            read_mode = locals().get('read_mode', "raw lines (treat each line as full text)")
            if read_mode == "raw lines (treat each line as full text)":
                rows_iter, approx_rows = read_rows_raw_lines(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto"), estimate_row_count_from_bytes(file_bytes)
            elif read_mode == "csv columns (streaming, recommended)":
                rows_iter, approx_rows = iter_csv_selected_columns(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto", delimiter, has_header, selected_cols, join_with), estimate_row_count_from_bytes(file_bytes)
            else:
                usecols = selected_cols if has_header else [int(c.split('_')[-1]) for c in selected_cols] if selected_cols else None
                rows_iter, approx_rows = read_rows_csv_pandas(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto", int(chunksize), has_header, usecols, join_with), estimate_row_count_from_bytes(file_bytes)
        elif is_xlsx and openpyxl:
            sheet = st.session_state.get(f"xlsx_sheet_{idx}", get_excel_sheetnames(file_bytes)[0] if get_excel_sheetnames(file_bytes) else None)
            if sheet:
                rows_iter = iter_excel_selected_columns(file_bytes, sheet, st.session_state.get(f"xlsx_has_header_{idx}", True), st.session_state.get(f"xlsx_cols_{idx}", []), st.session_state.get(f"xlsx_join_{idx}", " "))
                approx_rows = excel_estimate_rows(file_bytes, sheet, st.session_state.get(f"xlsx_has_header_{idx}", True))
        else: rows_iter, approx_rows = read_rows_raw_lines(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto"), estimate_row_count_from_bytes(file_bytes)
        
        if approx_rows <= 50_000: update_every = 500
        elif approx_rows <= 500_000: update_every = 2_000
        else: update_every = 10_000
        start_wall = time.perf_counter()
        
        def make_progress_cb(total_hint: int):
            def _cb(done: int):
                elapsed = time.perf_counter() - start_wall
                if total_hint > 0:
                    pct = min(99, int(done * 100 / total_hint))
                    per_file_bar.progress(pct)
                    rps = (done / elapsed) if elapsed > 0 else 0
                    eta = (total_hint - done) / rps if rps > 0 else None
                    per_file_status.markdown(f"• rows: {done:,}/{total_hint:,} • elapsed: {format_duration(elapsed)} {'• ETA:'+format_duration(eta) if eta else ''} • speed: {rps:,.0f} r/s")
                else:
                    rps = (done / elapsed) if elapsed > 0 else 0
                    per_file_status.markdown(f"• rows: {done:,} • elapsed: {format_duration(elapsed)} • speed: {rps:,.0f} r/s")
            return _cb
        
        data = process_rows_iter(
            rows_iter=rows_iter,
            remove_chat_artifacts=remove_chat_artifacts,
            remove_html_tags=remove_html_tags, unescape_entities=unescape_entities, remove_urls=remove_urls,
            keep_hyphens=keep_hyphens, keep_apostrophes=keep_apostrophes,
            user_phrase_stopwords=tuple(user_phrase_stopwords), user_single_stopwords=tuple(user_single_stopwords),
            add_preps=add_preps, drop_integers=drop_integers, min_word_len=min_word_len,
            compute_bigrams=compute_bigrams, progress_cb=make_progress_cb(approx_rows), update_every=update_every,
        )

        per_file_bar.progress(100)
        per_file_status.markdown(f"✅ done in {format_duration(time.perf_counter() - start_wall)} • rows: {data['rows']:,}")
        file_results.append({"name": fname, "counts": data["counts"], "font_path": per_file_font_path, "font_name": per_file_font_choice})
        combined_counts.update(data["counts"])
        if compute_bigrams: combined_bigrams.update(data["bigrams"])

        if data["counts"]:
            render_start = time.perf_counter()
            color_func = None
            if enable_sentiment:
                sentiments = get_sentiments(analyzer, tuple(data["counts"].keys()))
                color_func = create_sentiment_color_func(sentiments, pos_color, neg_color, neu_color, pos_threshold, neg_threshold)
            fig, _ = build_wordcloud_figure_from_counts(data["counts"], max_words, width, height, bg_color, colormap, per_file_font_path, random_state, color_func)
            col1, col2 = st.columns([3, 1])
            with col1: st.pyplot(fig, use_container_width=True)
            with col2:
                st.markdown(f"font: {per_file_font_choice}")
                st.caption(f"rendered in {format_duration(time.perf_counter() - render_start)}")
                st.download_button(f"📥 download {fname} word cloud", fig_to_png_bytes(fig), f"{fname}_wordcloud.png", "image/png")
            plt.close(fig); gc.collect()
        else: st.warning(f"no tokens remaining after filtering for {fname}.")

        overall_bar.progress(int(((idx + 1) / total_files) * 100))
        overall_status.markdown(f"processed {idx + 1} / {total_files} files")

# ---------------------------
# combined results
# ---------------------------
term_sentiments = {}
if enable_sentiment and combined_counts:
    term_sentiments = get_sentiments(analyzer, tuple(combined_counts.keys()))
    if compute_bigrams:
        bigram_phrases = tuple(" ".join(bg) for bg in combined_bigrams.keys())
        bigram_sentiments = get_sentiments(analyzer, bigram_phrases)
        term_sentiments.update(bigram_sentiments)
st.subheader("🖼️ combined word cloud")
if combined_counts:
    try:
        render_start = time.perf_counter()
        combined_color_func = None
        if enable_sentiment: combined_color_func = create_sentiment_color_func(term_sentiments, pos_color, neg_color, neu_color, pos_threshold, neg_threshold)
        fig, _ = build_wordcloud_figure_from_counts(combined_counts, max_words, width, height, bg_color, colormap, combined_font_path, random_state, combined_color_func)
        st.pyplot(fig, use_container_width=True)
        st.caption(f"rendered in {format_duration(time.perf_counter() - render_start)}")
        st.download_button("📥 download combined word cloud", fig_to_png_bytes(fig), "combined_wordcloud.png", "image/png")
        plt.close(fig); gc.collect()
    except MemoryError: st.error("memoryerror: try reducing max words or image size.")
else: st.info("upload files to generate word clouds and frequency tables.")
if combined_counts:
    if enable_sentiment:
        total_score, total_weight = 0.0, 0
        for word, freq in combined_counts.items():
            score = term_sentiments.get(word, 0.0)
            total_score += score * freq
            total_weight += freq
        avg_sentiment = total_score / total_weight if total_weight > 0 else 0.0
        sentiment_emoji = "😊" if avg_sentiment >= pos_threshold else ("😠" if avg_sentiment <= neg_threshold else "😐")
        st.metric(label=f"overall corpus sentiment {sentiment_emoji}", value=f"{avg_sentiment:.3f}", help="weighted average sentiment; higher is more positive")
    st.subheader(f"📊 combined frequency table (top {top_n} unigrams)")
    most_common_words = combined_counts.most_common(top_n)
    if enable_sentiment:
        freq_data = [[word, freq, term_sentiments.get(word, 0.0), get_sentiment_category(term_sentiments.get(word, 0.0), pos_threshold, neg_threshold)] for word, freq in most_common_words]
        freq_df = pd.DataFrame(freq_data, columns=["word", "frequency", "sentiment_score", "sentiment_category"])
    else: freq_df = pd.DataFrame(most_common_words, columns=["word", "frequency"])
    st.dataframe(freq_df, use_container_width=True)
    st.download_button("⬇️ download unigram frequencies (csv)", freq_df.to_csv(index=False).encode("utf-8"), "unigrams_top.csv", "text/csv")
    if compute_bigrams:
        st.subheader(f"📊 combined frequency table (top {top_n} bigrams)")
        if combined_bigrams:
            top_bigrams = combined_bigrams.most_common(top_n)
            if enable_sentiment:
                bigram_data = [[" ".join(bg), freq, term_sentiments.get(" ".join(bg), 0.0), get_sentiment_category(term_sentiments.get(" ".join(bg), 0.0), pos_threshold, neg_threshold)] for bg, freq in top_bigrams]
                bigram_df = pd.DataFrame(bigram_data, columns=["bigram", "frequency", "sentiment_score", "sentiment_category"])
            else:
                bigram_rows = [(" ".join(bg), freq) for bg, freq in top_bigrams]
                bigram_df = pd.DataFrame(bigram_rows, columns=["bigram", "frequency"])
            st.dataframe(bigram_df, use_container_width=True)
            st.download_button("⬇️ download bigram frequencies (csv)", bigram_df.to_csv(index=False).encode("utf-8"), "bigrams_top.csv", "text/csv")
        else: st.info("no bigrams found after filtering.")
    else: st.caption("bigrams are disabled. enable them in ⚙️ performance options")

# ---------------------------
# help / guide
# ---------------------------
with st.expander("ℹ️ how to use this app", expanded=False):
    # V23: Added privacy warning to help section
    st.error("**Important:** Never upload files with private or sensitive information. All data is processed on public servers.")
    
    st.markdown("""
    #### IF NEEDED, Here's HOW to Prepare Your Data Securely
    Before uploading, it is your responsibility to remove or mask all Personally Identifiable Information (PII). Here are some recommended methods:

    *   **Remove Columns (Easiest):** In your spreadsheet program, simply delete any columns containing names, email addresses, phone numbers, etc.
    *   **Pseudonymize:** Use find-and-replace to substitute real names with generic identifiers (e.g., replace all instances of "John Smith" with "Speaker 1").
    *   **For Advanced Users (Scripting):** A local Python script can be used to one-way hash sensitive IDs. For example, replace a `user_id` column with a `hashed_user_id` column using a library like `hashlib`.

    ---
    #### Application Guide
    - **Upload:** Add one or more CSV, Excel, or VTT files.
    - **Configure Input (per-file):** Use the expander for each file to set up how data is read.
    - **Customize Appearance:** Adjust colors, fonts, and image sizes for the word clouds.
    - **Sentiment Analysis:** Enable this to color-code word clouds and add sentiment data to the tables.
    - **Cleaning & Stopwords:**
      - Use the "remove chat/transcript artifacts" checkbox for text from Slack, Zoom, and MS Teams.
      - Fine-tune what gets included by removing HTML, URLs, short words, and custom stopwords (like speaker names).
    - **Performance:** This app is designed for large files using streaming. Bi-gram computation is optional.
    - **Outputs:** Download word clouds and CSV files for top unigrams and bi-grams.

    """)


